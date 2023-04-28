from openpyxl import load_workbook
#оформить в тест, добавить ассерты и использовать универсальный путь
from tests.funtions import resources_path

def test_xlsx_info():
    file = resources_path('resources', 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(file)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert sheet.cell(row=3, column=2).value == 'Mara'
